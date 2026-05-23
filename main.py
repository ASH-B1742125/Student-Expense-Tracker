import openpyxl as xl
#For  Charts
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

#To Load The .xlsx since it returs obj need to store
wb = xl.load_workbook('student_expense.xlsx')
#create a sheet
sheet = wb.active

total_spent =0
category_total = {}
#Access all expenses from excel except the header
for row in range(2 , sheet.max_row + 1) :
   #this store the value of each header
   date = sheet.cell(row, 1).value
   item = sheet.cell(row, 2).value
   category = sheet.cell(row, 3).value
   amount = sheet.cell(row, 4).value
   
   print(date, item, category, amount)

   #add each expense to total
   total_spent += amount

   # category-wise expense
   if category in category_total :
      category_total[category] += amount
   else :
      category_total[category] = amount

print()
print("----- CATEGORY BREAKDOWN -----")
for cat,amt in category_total.items():
   print(f"{cat} -> {amt}")

print()
print(f"The Total is: {total_spent}")

# creating new sheet
report_workbook = xl.Workbook()
report_sheet = report_workbook.active

#add title
report_sheet["A1"] = "Expense Report"

#total spending
report_sheet["A2"] = "Total Spent"
report_sheet["B2"] = total_spent

#category breakdown
report_sheet["A4"] = "CATEGORY BREAKDOWN"
row = 5

for cat , amt in category_total.items() :
   report_sheet.cell(row,1).value = cat
   report_sheet.cell(row, 2).value = amt
   row += 1

#creating pie chart
pie = PieChart()
#creating category data reference
data = Reference(
    report_sheet,
    min_col=2,
    min_row=5,
    max_row=row-1
)

#category labels reference
labels = Reference(
    report_sheet,
    min_col=1,
    min_row=5,
    max_row=row-1
)

#connecting the amount data to the pie chart
pie.add_data(data)

#attaching the category labels to the pie chart
pie.set_categories(labels)

#chart title
pie.title = "Expense Distribution"

#Placeing chart in worksheet
report_sheet.add_chart(pie, "D2")


#creating Bar Chart
bar = BarChart()
bar.add_data(data)
bar.set_categories(labels)

#Adding Titles
bar.title = "Expense by Category"
bar.x_axis.title = "Categories"
bar.y_axis.title = "Amount"

report_sheet.add_chart(bar, "D20")

print("Sheet Is Created!")

#save the file
report_workbook.save("expense_report.xlsx")
